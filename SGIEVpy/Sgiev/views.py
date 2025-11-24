from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import models, transaction
from django.http import JsonResponse
from datetime import datetime, date, timedelta
from decimal import Decimal, InvalidOperation
from .models import (
    Categoria, 
    Proveedor,
    Usuarios, 
    Producto, 
    Venta, 
    Venta_has_producto, 
    Movimiento_inventario,
    Envio, 
    Mensajeria,
    Compra_proveedor,
    Compra_detalle
)
from .forms import LoginForm, UsuarioForm, VentaForm, AgregarProductoForm,EditarEstadoVentaForm,EnvioForm, MensajeriaForm,EnvioEditarOperarioForm
from .decorators import admin_required
import json
from django.utils.safestring import mark_safe
from django.template.loader import render_to_string
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db.models import Q
from django.db.models import F 
from django.template.loader import get_template
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from django.views.decorators.cache import never_cache
from django.http import HttpResponseRedirect 
from django.urls import reverse
from django.utils import timezone
from django.db import transaction
from uuid import uuid4
from django.db.models import Sum, Count, Q, F
from django.utils import timezone
from datetime import datetime, date, timedelta
from decimal import Decimal
import json

def index(request):  
    """
    Vista principal - Landing page
    """
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    return render(request, 'index.html')


def login(request):
    return render(request, 'login.html')


def admin(request):
    return render(request, 'admin.html')

#dashboard
@login_required(login_url='login')
def dashboard_view(request):
    """
    Vista principal del dashboard con estadísticas reales
    """
    from django.db.models import Sum, Count
    from django.utils import timezone
    from datetime import timedelta
    import json
    
    usuario = request.user
    es_admin = usuario.tipo_usu == 'administrador'
    
    # Fecha actual y rango del mes
    hoy = timezone.now()
    inicio_mes = hoy.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # ===== FILTRAR DATOS SEGÚN EL ROL =====
    if es_admin:
        ventas_query = Venta.objects.all()
        envios_query = Envio.objects.all()
    else:
        ventas_query = Venta.objects.filter(usuarios_id_usuario=usuario)
        envios_query = Envio.objects.filter(usuarios_id_usuario=usuario)
    
    # ===== ESTADÍSTICAS DE VENTAS =====
    
    # Ingresos del mes actual
    ventas_mes = ventas_query.filter(
        fecha_factura__gte=inicio_mes,
        estado_pago__in=['pagado', 'parcial']
    )
    
    ingresos_mes = ventas_mes.aggregate(total=Sum('valor_total'))['total'] or Decimal('0')
    
    # Ventas totales (histórico)
    ventas_totales = ventas_query.filter(
        estado_pago__in=['pagado', 'parcial']
    ).aggregate(total=Sum('valor_total'))['total'] or Decimal('0')
    
    # Productos vendidos (cantidad total de items)
    try:
        productos_vendidos = Venta_has_producto.objects.filter(
            venta_idfactura__in=ventas_query,
            venta_idfactura__fecha_factura__gte=inicio_mes
        ).aggregate(total=Sum('cantidad'))['total'] or 0
    except Exception as e:
        productos_vendidos = 0
    
    # ===== ESTADÍSTICAS DE ENVÍOS =====
    
    # Envíos pendientes
    envios_pendientes = envios_query.filter(estado_envio='pendiente').count()
    
    # Envíos en tránsito
    envios_transito = envios_query.filter(estado_envio='en_transito').count()
    
    # Envíos entregados este mes
    envios_entregados_mes = envios_query.filter(
        estado_envio='entregado',
        fecha_entrega__gte=inicio_mes
    ).count()
    
    # Envíos devueltos este mes (AGREGAR ESTO)
    envios_devueltos_mes = envios_query.filter(
        estado_envio='devuelto',
        fecha_entrega__gte=inicio_mes
    ).count()
    
    # Últimos envíos entregados (AGREGAR ESTO)
    envios_entregados_recientes = envios_query.filter(
        estado_envio='entregado'
    ).select_related('venta_idfactura', 'fk_mensajeria').order_by('-fecha_entrega')[:5]
    
    # ===== ALERTAS DE INVENTARIO (SOLO ADMIN) =====
    
    if es_admin:
        productos_stock_bajo = Producto.objects.filter(
            activo=1,
            stock_actual__lte=models.F('stock_minimo')
        ).order_by('stock_actual')[:10]
    else:
        productos_stock_bajo = []
    
    # ===== ÚLTIMAS VENTAS =====
    
    ventas_recientes = list(ventas_query.order_by('-fecha_factura')[:5])
    historial_ventas = list(ventas_query.order_by('-fecha_factura')[:10])
    
    # ===== DATOS PARA EL GRÁFICO =====
    
    meses_labels = []
    meses_valores = []
    
    meses = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
    
    for i in range(5, -1, -1):
        fecha_mes = hoy - timedelta(days=30 * i)
        inicio = fecha_mes.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        if i == 0:
            fin = hoy
        else:
            siguiente_mes = inicio + timedelta(days=32)
            fin = siguiente_mes.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        total_mes = ventas_query.filter(
            fecha_factura__gte=inicio,
            fecha_factura__lt=fin,
            estado_pago__in=['pagado', 'parcial']
        ).aggregate(total=Sum('valor_total'))['total'] or Decimal('0')
        
        mes_nombre = meses[fecha_mes.month - 1]
        meses_labels.append(mes_nombre)
        meses_valores.append(float(total_mes))
    
    # ===== PORCENTAJES =====
    
    inicio_mes_anterior = (inicio_mes - timedelta(days=1)).replace(day=1)
    ventas_mes_anterior = ventas_query.filter(
        fecha_factura__gte=inicio_mes_anterior,
        fecha_factura__lt=inicio_mes,
        estado_pago__in=['pagado', 'parcial']
    ).aggregate(total=Sum('valor_total'))['total'] or Decimal('0')
    
    if ventas_mes_anterior > 0:
        ingresos_porcentaje = round(((ingresos_mes - ventas_mes_anterior) / ventas_mes_anterior) * 100, 1)
    else:
        ingresos_porcentaje = 100 if ingresos_mes > 0 else 0
    
    # ===== ENVÍOS PRÓXIMOS =====
    
    envios_proximos = list(envios_query.filter(
        estado_envio='en_transito',
        fecha_entrega__lte=hoy + timedelta(days=3)
    ).order_by('fecha_entrega')[:5])
    
    # ===== CONTEXTO =====
    
    context = {
        'usuario': usuario,
        'es_admin': es_admin,
        
        # Estadísticas principales
        'estadisticas_dashboard': {
            'ingresos_mes': ingresos_mes,
            'ingresos_porcentaje': abs(ingresos_porcentaje),
            'ingresos_crecimiento': ingresos_porcentaje >= 0,
            'ventas_totales': ventas_totales,
            'productos_vendidos': productos_vendidos,
            'envios_pendientes': envios_pendientes,
            'envios_transito': envios_transito,
            'envios_entregados': envios_entregados_mes,
            'envios_devueltos': envios_devueltos_mes,
        },
        
        # Ventas
        'ventas_recientes': ventas_recientes,
        'historial_ventas': historial_ventas,
        
        # Gráfico
        'grafico_labels': json.dumps(meses_labels),
        'grafico_valores': json.dumps(meses_valores),
        
        # Alertas
        'productos_stock_bajo': productos_stock_bajo,
        'envios_proximos': envios_proximos,
        'envios_entregados_recientes': list(envios_entregados_recientes),
        'tiene_alertas': len(productos_stock_bajo) > 0 or envios_pendientes > 0,
    }
    
    return render(request, 'dashboard.html', context)

# CATEGORIA

def inicio_cat(request):
    return render(request, 'categoria/index.html')


from django.core.paginator import Paginator
from django.shortcuts import render
from .models import Categoria

def list_categoria(request):
    search = request.GET.get('search', '')

    categorias = Categoria.objects.all()

    # Filtro por búsqueda
    if search:
        categorias = categorias.filter(
            nombre_categoria__icontains=search
        )

    # Paginación (10 por página)
    paginator = Paginator(categorias, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    usuario = request.user                     # <-- usuario logueado
    es_admin = request.user.tipo_usu == 'administrador'  # <-- bandera admin

    context = {
        'page_obj': page_obj,
        'search': search,
        'usuario': usuario,     # <-- añadido para badge
        'es_admin': es_admin    # <-- añadido para badge
    }

    return render(request, 'categoria/index.html', context)




def registro_categoria(request):
    if request.method == "POST":
        nombre = request.POST.get('nombreCat')
        descripcion = request.POST.get('descCat')
        fecha = datetime.now()
        estado = 1

        categoria = Categoria(
            nombre_categoria=nombre,
            descripcion_categoria=descripcion,
            fecha_creacion=fecha,
            activo=estado
        )
        
        categoria.save()
        return redirect('list_categoria')


    usuario = request.user                     # <-- usuario logueado
    es_admin = request.user.tipo_usu == 'administrador'  # <-- bandera del rol

    return render(request, 'categoria/nuevocat.html', {
        'usuario': usuario,      # <-- añadido para badge
        'es_admin': es_admin     # <-- añadido para badge
    })


def pre_editar_categoria(request, id):
    categoria = Categoria.objects.get(id=id)
    data = {
        'categoria': categoria
    }
    return render(request, 'categoria/editarcat.html', data)


def editar_categoria(request, id):
    if request.method == "POST":
        categoria = Categoria.objects.get(id=id)

        nombre = request.POST.get('nombreCat')
        descripcion = request.POST.get('descCat')
        estado = request.POST.get('estadoCat')

        categoria.nombre_categoria = nombre
        categoria.descripcion_categoria = descripcion
        categoria.activo = estado

        categoria.save()
    return redirect("categoria/index")


def eliminar_categoria(request, id):
    categoria = Categoria.objects.get(id=id)
    categoria.delete()
    return redirect('list_categoria')

#PRODUCTOS
@never_cache
def list_producto(request):
    

    producto = Producto.objects.select_related('categoria_idcategoria', 'proveedor_idproveedor').all()
    

    lotes_disponibles = []
    for p in producto:
        nombre = p.nombre_producto.strip()
        descripcion = p.descripcion_producto.strip()
        nombre_completo = f"{nombre} {descripcion}" 
        
        lotes_disponibles.append({
            'id': p.id, 
            'nombre_producto_completo': nombre_completo, 
            'codigo_barras': p.codigo_barras, 
            'stock_actual': p.stock_actual, 
            'fecha_vencimiento': p.fecha_vencimiento.strftime('%Y-%m-%d') if p.fecha_vencimiento else 'N/A'
        })

    lotes_json_string = json.dumps(lotes_disponibles)
    lotes_json_safe = mark_safe(lotes_json_string) 
    
    
 
    try:
        historial_compras = Compra_proveedor.objects.select_related(
        'producto_idproducto__proveedor_idproveedor', 
        'usuarios_id_usuario'
    ).all().order_by('-fecha_compra')[:10]
    except:
        print(f"Error en select_related de Compra_proveedor: {e}")
    historial_compras = Compra_proveedor.objects.all().order_by('-fecha_compra')[:10]


    try:
        historial_salidas = Movimiento_inventario.objects.filter(
            tipo_movimiento__in=['ajuste', 'venta']
        ).order_by('-fecha_movimiento')[:15]
    except:
        print(f"Error al cargar historial de salidas: {e}")
        historial_salidas = [] 

    data = {
        'producto': producto,
        'lotes_json': lotes_json_safe,
        'historial_compras': historial_compras,  
        'historial_salidas': historial_salidas, 
    }
    

    return render(request, 'producto/index.html', data)


def detalle_producto_modal(request, producto_id):
    producto = get_object_or_404(Producto, pk=producto_id)
    

    movimientos_producto = Movimiento_inventario.objects.filter(
        producto_idproducto=producto
    ).order_by('-fecha_movimiento')[:5] 
    
    context = {
        'p': producto,
        'movimientos': movimientos_producto,
    }
 
    return render(request, 'producto/detalle_producto_modal_content.html', context)

def registro_producto(request):
    if request.method == "POST":
        
        nombre = request.POST.get('nombre_producto')
        descripcion = request.POST.get('descripcion_producto')
        codigo = request.POST.get('codigo_barras')
        registro_sanitario = request.POST.get('registrosanitario')
        precio_compra = request.POST.get('precio_compra') or "0"
        precio_venta = request.POST.get('precio_venta') or "0"
        margen = request.POST.get('margen_ganancia') or "0"
        stock_actual = request.POST.get('stock_actual') or "0"
        stock_min = request.POST.get('stock_minimo') or "0"
        stock_max = request.POST.get('stock_maximo') or "0"
        fecha_ven = request.POST.get('fecha_vencimiento')
        fecha_cre = datetime.now()
        categoria_id = request.POST.get('categoria_idcategoria')
        proveedor_id = request.POST.get('proveedor_idproveedor')
        activo = request.POST.get('activo') or 1

        
        try:
            precio_compra = Decimal(precio_compra)
        except InvalidOperation:
            precio_compra = Decimal('0.00')
        try:
            precio_venta = Decimal(precio_venta)
        except InvalidOperation:
            precio_venta = Decimal('0.00')
        try:
            margen = Decimal(margen)
        except InvalidOperation:
            margen = Decimal('0.00')

        try:
            stock_actual = int(stock_actual)
        except:
            stock_actual = 0
        try:
            stock_min = int(stock_min)
        except:
            stock_min = 0
        try:
            stock_max = int(stock_max)
        except:
            stock_max = 0

       
        categoria = get_object_or_404(Categoria, id=categoria_id)
        proveedor = get_object_or_404(Proveedor, id=proveedor_id)

        producto = Producto(
            nombre_producto = nombre,
            descripcion_producto = descripcion,
            codigo_barras = codigo,
            registrosaniario = registro_sanitario,
            precio_compra = precio_compra,
            precio_venta = precio_venta,
            margen_ganancia = margen,
            stock_actual = stock_actual,
            stock_minimo = stock_min,
            stock_maximo = stock_max,
            fecha_vencimiento = fecha_ven,
            categoria_idcategoria = categoria,
            proveedor_idproveedor = proveedor,
            activo = activo
        )
        producto.save()
        return redirect('list_producto')

  
    categorias = Categoria.objects.filter(activo=1)
    proveedores = Proveedor.objects.filter(activo=1)
    data = {'categorias': categorias, 'proveedores': proveedores}
    return render(request, 'producto/nuevoprod.html', data)



def pre_editar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    categorias = Categoria.objects.filter(activo=1)
    proveedores = Proveedor.objects.filter(activo=1)
    data = {
        'producto': producto,
        'categorias': categorias,
        'proveedores': proveedores
    }
    return render(request, 'producto/editarprod.html', data)



def editar_producto(request, id):
    if request.method == "POST":
        producto = get_object_or_404(Producto, id=id)

        producto.nombre_producto = request.POST.get('nombre_producto')
        producto.descripcion_producto = request.POST.get('descripcion_producto')
        producto.codigo_barras = request.POST.get('codigo_barras')
        producto.registrosaniario = request.POST.get('registrosanitario')

        
        from decimal import Decimal, InvalidOperation
        try:
            producto.precio_compra = Decimal(request.POST.get('precio_compra') or '0')
        except InvalidOperation:
            producto.precio_compra = Decimal('0.00')
        try:
            producto.precio_venta = Decimal(request.POST.get('precio_venta') or '0')
        except InvalidOperation:
            producto.precio_venta = Decimal('0.00')
        try:
            producto.margen_ganancia = Decimal(request.POST.get('margen_ganancia') or '0')
        except InvalidOperation:
            producto.margen_ganancia = Decimal('0.00')

        try:
            producto.stock_actual = int(request.POST.get('stock_actual') or 0)
        except:
            producto.stock_actual = 0
        try:
            producto.stock_minimo = int(request.POST.get('stock_minimo') or 0)
        except:
            producto.stock_minimo = 0
        try:
            producto.stock_maximo = int(request.POST.get('stock_maximo') or 0)
        except:
            producto.stock_maximo = 0

        producto.fecha_vencimiento = request.POST.get('fecha_vencimiento')

        
        categoria_id = request.POST.get('categoria_idcategoria')
        proveedor_id = request.POST.get('proveedor_idproveedor')
        producto.categoria_idcategoria = get_object_or_404(Categoria, id=categoria_id)
        producto.proveedor_idproveedor = get_object_or_404(Proveedor, id=proveedor_id)

        producto.activo = request.POST.get('activo') or producto.activo

        producto.save()

    return redirect('list_producto')


@transaction.atomic
def eliminar_producto(request, id):
    
    usuario_id = request.session.get('_auth_user_id') 
    usuario_actual = None
    
    if usuario_id:
        try:
            usuario_actual = Usuarios.objects.get(id=usuario_id)
        except Usuarios.DoesNotExist:
            pass 
            
 
    producto = get_object_or_404(Producto.objects.select_for_update(), id=id)

    cantidad_eliminada = producto.stock_actual
    
    Movimiento_inventario.objects.create(
        producto_idproducto=producto,
        cantidad=cantidad_eliminada,
        tipo_movimiento='ajuste',   
        stock_anterior=producto.stock_actual,
        stock_nuevo=0, 
        
        precio_unitario=producto.precio_compra,
        valor_total=producto.precio_compra * cantidad_eliminada,
        
        referencia_id=producto.id, 
        tipo_referencia='ajuste', 
        
        observaciones=f'AJUSTE NEGATIVO: Eliminación total. Usuario ID: {usuario_id if usuario_id else "No identificado"}',
        
        usuarios_id_usuario=usuario_actual, 
        fecha_movimiento=timezone.now()
    )
    
    producto.delete()
    
    return redirect('list_producto')

def generar_reporte_productos(request):
    
    categoria_id = request.GET.get('categoria')
    stock_estado = request.GET.get('stock_estado')
    formato = request.GET.get('formato')

    productos_query = Producto.objects.all().order_by('nombre_producto')
    
    productos_query = productos_query.select_related('categoria_idcategoria')

   
    if categoria_id:
    
        if categoria_id.isdigit():
            productos_query = productos_query.filter(categoria_idcategoria__id=categoria_id)

    if stock_estado == 'bajo':
       
        productos_query = productos_query.filter(stock_actual__lte=F('stock_minimo'))
        
    elif stock_estado == 'vencido':
       
        fecha_limite = date.today() + timedelta(days=30)
        productos_query = productos_query.filter(fecha_vencimiento__lte=fecha_limite).order_by('fecha_vencimiento')


 
    if formato == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="reporte_inventario_filtrado.xlsx"'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Inventario Romar Natural"

        
        columns = [
            'ID', 'Nombre', 'Descripción', 'Lote', 'Categoría', 
            'Stock Actual', 'Stock Mínimo', 'Precio Venta', 'Vencimiento'
        ]
        
       
        row_num = 1
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            
     
        for producto in productos_query:
            row_num += 1
            row = [
                producto.id,
                producto.nombre_producto,
                producto.descripcion_producto,
                producto.codigo_barras,
                producto.categoria_idcategoria.nombre_categoria if producto.categoria_idcategoria else 'Sin Categoría',
                producto.stock_actual,
                producto.stock_minimo,
                producto.precio_venta,
                producto.fecha_vencimiento.strftime("%Y-%m-%d") if producto.fecha_vencimiento else 'N/A'
            ]
            
            for col_num, cell_value in enumerate(row, 1):
                worksheet.cell(row=row_num, column=col_num, value=cell_value)
        
        workbook.save(response)
        return response


    elif formato == 'pdf':
        
       
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_inventario_filtrado.pdf"'

       
        doc = SimpleDocTemplate(response, pagesize=letter)
        styles = getSampleStyleSheet()
        story = [] 
        story.append(Paragraph("<b>REPORTE DE INVENTARIO - ROMAR NATURAL</b>", styles['h1']))
        story.append(Paragraph(f"Fecha del Reporte: {date.today().strftime('%Y-%m-%d')}", styles['Normal']))
        
        filtro_desc = "Todos los productos"
        if stock_estado == 'bajo':
            filtro_desc = "Productos con Stock Bajo/Mínimo"
        elif stock_estado == 'vencido':
            filtro_desc = "Productos Próximos a Vencer (30 días)"
            
        story.append(Paragraph(f"Filtro Aplicado: {filtro_desc}", styles['Normal']))
        story.append(Paragraph("<br/>", styles['Normal'])) 
        
      
        data = [
            ['ID', 'Nombre Producto', 'Lote', 'Stock', 'Mínimo', 'Vencimiento', 'Categoría']
        ]
        
        for producto in productos_query:
            data.append([
                producto.id,
                f"{producto.nombre_producto} {producto.descripcion_producto}",
                producto.codigo_barras,
                producto.stock_actual,
                producto.stock_minimo,
                producto.fecha_vencimiento.strftime("%Y-%m-%d") if producto.fecha_vencimiento else 'N/A',
                producto.categoria_idcategoria.nombre_categoria if producto.categoria_idcategoria else 'S/C'
            ])

        
        if not data or len(data) == 1:
            story.append(Paragraph("No se encontraron productos con los filtros seleccionados.", styles['Normal']))
        else:
            table = Table(data, colWidths=[40, 150, 80, 50, 50, 90, 100])
            
            
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
              
                ('BACKGROUND', (0, 0), (-1, -1), colors.white),
            ]))
            story.append(table)

        
        doc.build(story)
        
        return response
        

    return HttpResponse("Formato de reporte no válido.", status=400)


def registrar_salida_inventario_ajuste(request):
    if request.method == "POST":
        
        producto_id_lote = request.POST.get('producto_id_lote') 
        codigo_barras_lote = request.POST.get('codigo_barras_hidden') 
        cantidad_salida = request.POST.get('cantidad_salida')
        motivo_salida = request.POST.get('motivo_salida')

        
        if not producto_id_lote or not producto_id_lote.isdigit():
            messages.error(request, "Error: Producto no seleccionado o ID no válido.")
            return redirect('list_producto')

        try:
            cantidad = int(cantidad_salida)
        except ValueError:
            messages.error(request, "Error: Cantidad inválida.")
            return redirect('list_producto')

        
    
        response_redirect = HttpResponseRedirect(reverse('list_producto'))
        response_redirect['Cache-Control'] = 'no-cache, no-store, must-revalidate' 
    
        try:
            with transaction.atomic():
                
               
                producto = Producto.objects.select_for_update().get(id=producto_id_lote)
                
               
                usuario = None
                if request.user.is_authenticated:
                    try:
                        usuario = Usuarios.objects.get(id=request.user.id) 
                    except Usuarios.DoesNotExist:
                        
                        messages.error(request, "Error de autenticación: El usuario logueado no está registrado en el modelo de Usuarios.")
                        raise Exception("Usuario no encontrado en el modelo Usuarios.") 
                else:
                    
                    messages.error(request, "Error de autenticación: Debe iniciar sesión para registrar movimientos.")
                    raise Exception("Usuario no autenticado.")


                
                if cantidad <= 0 or cantidad > producto.stock_actual:
                    messages.error(request, f"Error: Stock insuficiente. Stock actual del lote {producto.codigo_barras}: {producto.stock_actual}")
                    raise Exception("Stock Insuficiente.")
                
                
                stock_anterior = producto.stock_actual
                precio_unitario = producto.precio_compra 
                valor_total = precio_unitario * cantidad

                Producto.objects.filter(id=producto_id_lote).update(stock_actual=F('stock_actual') - cantidad)

                producto.refresh_from_db()

                stock_final_real = producto.stock_actual
                print(f"DEBUG POST-UPDATE: Stock verificado en DB: {stock_final_real}")
                

                movimiento = Movimiento_inventario(
                    producto_idproducto=producto, 
                    usuarios_id_usuario=usuario, 
                    
                    tipo_movimiento='ajuste', 
                    cantidad=cantidad,
                    stock_anterior=stock_anterior,
                    stock_nuevo=stock_final_real, 
                    precio_unitario=precio_unitario,
                    valor_total=valor_total,
                    
                    referencia_id=0, 
                    tipo_referencia='ajuste', 
                    imagen_comprobante='',
                    
                    observaciones=f"AJUSTE POR BAJA/DEVOLUCIÓN. Motivo: {motivo_salida}. Lote Retirado (Código de Barras): {codigo_barras_lote}", 
                )
                movimiento.save()
                
            messages.success(request, f"Salida de {cantidad} unidades registrada. Nuevo stock: {stock_final_real}")
            return response_redirect 
            
        except Exception as e:
            
            
            if str(e) not in ["Stock Insuficiente.", "Usuario no encontrado en el modelo Usuarios.", "Usuario no autenticado."]:
                 print(f"ERROR FATAL (Rollback): Falló al registrar el ajuste de inventario. CAUSA: {e}")
                 messages.error(request, "Error fatal al procesar el ajuste. El stock no fue modificado.")
            
            return response_redirect 

    return redirect('list_producto')

#VENTA DE PROVEEDOR
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction

@login_required 
@transaction.atomic
def crear_compra_proveedor(request, idproveedor):
    proveedor = get_object_or_404(Proveedor, id=idproveedor)
    productos = Producto.objects.filter(activo=1)
    categorias = Categoria.objects.filter(activo=1)

    if request.method == "POST":
        
        tipo = request.POST.get("tipo_producto")
        estado = request.POST.get("estado_compra", "recibida")
        observaciones = request.POST.get("observaciones", "")
        numero_factura = request.POST.get("numero_factura", "")
        fecha_vencimiento_detalle = request.POST.get("fecha_vencimiento") or None
        lote = request.POST.get("lote", "")
        
        try:
            cantidad = int(request.POST.get("cantidad", "0"))
            precio_unitario = Decimal(request.POST.get("valor_unitario") or "0")
        except:
            messages.error(request, "Error en el formato de la cantidad o el valor unitario.")
         
            return redirect("listar_proveedores") 

        if cantidad <= 0 or precio_unitario <= Decimal('0'):
            messages.error(request, "La cantidad y el precio unitario deben ser mayores a cero.")
            return redirect("listar_proveedores")

        subtotal_linea = cantidad * precio_unitario
        iva = subtotal_linea * Decimal("0.19")
        total = subtotal_linea + iva

        producto = None
        
    
        if tipo == "nuevo":
            categoria_id = request.POST.get('categoria')
            nombre_producto = request.POST.get('nombre_producto')
            precio_venta_post = request.POST.get('precio_venta')

            if not categoria_id or not nombre_producto:
                 messages.error(request, "Nombre y Categoría son requeridos para el producto nuevo.")
                 return redirect("listar_proveedores")
            
           
            
            categoria_obj = get_object_or_404(Categoria, id=categoria_id)
            precio_venta = Decimal(precio_venta_post or '0')

            
            margen_calculado = Decimal('0.00')
            if precio_unitario > Decimal('0'):
                margen_calculado = ((precio_venta - precio_unitario) / precio_unitario) * Decimal('100.00')
                margen_calculado = margen_calculado.quantize(Decimal('0.01'))

            
            producto = Producto.objects.create(
                nombre_producto=nombre_producto,
                descripcion_producto=request.POST.get('descripcion_producto', ''),
                codigo_barras=request.POST.get('codigo_barras', ''),
                registrosaniario=request.POST.get('registrosaniario', ''),
                precio_compra=precio_unitario,
                precio_venta=precio_venta,
                margen_ganancia=margen_calculado,
                stock_actual=0, 
                stock_minimo=int(request.POST.get('stock_minimo') or 1),
                stock_maximo=int(request.POST.get('stock_maximo') or 1000),
                fecha_vencimiento=fecha_vencimiento_detalle, 
                categoria_idcategoria=categoria_obj, 
                proveedor_idproveedor=proveedor,
                activo=1
            )
        else:
            # PRODUCTO EXISTENTE
            producto_id = request.POST.get("producto_id")
            if not producto_id:
                messages.error(request, "Debe seleccionar un producto existente.")
                return redirect("listar_proveedores")
            producto = get_object_or_404(Producto, id=producto_id)

       # CREACIÓN DE LA COMPRA Y EL DETALLE 
        if not hasattr(request.user, 'pk'):
             messages.error(request, "Usuario no autenticado correctamente.")
             return redirect("login") 

        compra = Compra_proveedor.objects.create(
            numero_factura_compra=numero_factura if numero_factura else f"CMP{Compra_proveedor.objects.count()+1}",
            subtotal_compra=subtotal_linea,
            iva_compra=iva,
            total_compra=total,
            estado_compra=estado,
            observaciones_compra=observaciones,
            imagen_factura_compra="",
            usuarios_id_usuario=request.user,
            producto_idproducto=producto 
        )

        Compra_detalle.objects.create(
            compra_idcompra=compra,
            producto_idproducto=producto,
            cantidad=cantidad,
            precio_compra_unitario=precio_unitario,
            subtotal_linea_compra=subtotal_linea,
            lote=lote,
            fecha_vencimiento=fecha_vencimiento_detalle
        )

       
        producto.stock_actual += cantidad
        producto.precio_compra = precio_unitario 
        
        if producto.precio_venta > Decimal('0') and producto.precio_compra > Decimal('0'):
             nuevo_margen = ((producto.precio_venta - producto.precio_compra) / producto.precio_compra) * Decimal('100.00')
             producto.margen_ganancia = nuevo_margen.quantize(Decimal('0.01'))

        producto.save()

        messages.success(request, "Compra registrada exitosamente.")
        return redirect("listar_proveedores")

    return render(request, "proveedor/crear_compra.html", {
        "proveedor": proveedor,
        "productos": productos,
        "categorias": categorias,
    })




def agregar_al_carrito_compra(request, proveedor, productos, categorias):
    """
    Agrega un producto a la lista de compra (carrito) en la sesión.
    Se ha añadido la obligatoriedad del campo 'lote'.
    """
    
    tipo = request.POST.get("tipo_producto")
    
    try:
        cantidad = int(request.POST.get("cantidad", "0"))
        precio_unitario = Decimal(request.POST.get("valor_unitario") or "0")
        lote_detalle = request.POST.get("lote") 
        
        if not lote_detalle:
            messages.error(request, "El campo Lote es obligatorio para agregar al carrito.")
            return redirect('crear_compra_proveedor', idproveedor=proveedor.id)

        if cantidad <= 0 or precio_unitario <= Decimal('0'):
            messages.error(request, "La cantidad y el precio unitario deben ser mayores a cero.")
            return redirect('crear_compra_proveedor', idproveedor=proveedor.id)
            
    except:
        messages.error(request, "Error en el formato de la cantidad o el precio unitario.")
        return redirect('crear_compra_proveedor', idproveedor=proveedor.id) 

    fecha_vencimiento = request.POST.get("fecha_vencimiento") or None
    nombre_producto = ""
    producto_id = None
    
   
    if tipo == "nuevo":
        nombre_producto = request.POST.get('nombre_producto')
        categoria_id = request.POST.get('categoria')
        
        if not categoria_id or not nombre_producto:
            messages.error(request, "Nombre y Categoría son requeridos para el producto nuevo.")
            return redirect('crear_compra_proveedor', idproveedor=proveedor.id)

        if not fecha_vencimiento:
            messages.error(request, "La fecha de vencimiento es obligatoria para productos nuevos.")
            return redirect('crear_compra_proveedor', idproveedor=proveedor.id)    
        
        producto_id = None 
    else:
        
        producto_id_post = request.POST.get("producto_id")
        if not producto_id_post:
            messages.error(request, "Debe seleccionar un producto existente.")
            return redirect('crear_compra_proveedor', idproveedor=proveedor.id)
            
        
        producto_existente_obj = get_object_or_404(Producto, id=producto_id_post) 
        nombre_producto = producto_existente_obj.nombre_producto
        producto_id = producto_existente_obj.id
    
    
   
    carrito_key = f'carrito_compra_{proveedor.id}'
    carrito = request.session.get(carrito_key, [])
    
    subtotal_linea = cantidad * precio_unitario

   
    nuevo_item = {
        'temp_id': str(uuid4()), 
        'tipo': tipo,
        'producto_id': producto_id, 
        'nombre': nombre_producto,
        'precio': float(precio_unitario),
        'cantidad': cantidad,
        'subtotal': float(subtotal_linea),
        'fecha_vencimiento': fecha_vencimiento,
        'lote': lote_detalle, 
    }
    
  
    if tipo == "nuevo":
        nuevo_item.update({
            'categoria': categoria_id,
            'descripcion_producto': request.POST.get('descripcion_producto', ''),
            'registrosaniario': request.POST.get('registrosaniario', ''),
            'precio_venta': request.POST.get('precio_venta') or '0',
            'stock_minimo': request.POST.get('stock_minimo') or 1,
            'stock_maximo': request.POST.get('stock_maximo') or 1000,
            'codigo_barras_nuevo': request.POST.get('codigo_barras', ''), 
        })
    
   
    carrito.append(nuevo_item)
    
    
    request.session[carrito_key] = carrito
    request.session.modified = True
    
    messages.success(request, f'Producto añadido a la lista: {nombre_producto}')
    return redirect('crear_compra_proveedor', idproveedor=proveedor.id)


@transaction.atomic
def procesar_compra_final(request, proveedor):
    """
    Procesa la compra final, gestionando la creación/búsqueda de productos,
    actualización de stock y asegurando la integridad de los campos NOT NULL.
    """
    carrito_key = f'carrito_compra_{proveedor.id}'
    carrito = request.session.get(carrito_key, [])
    
    if not carrito:
        messages.error(request, 'La lista de compra está vacía.')
        return redirect('crear_compra_proveedor', idproveedor=proveedor.id) 
        
    estado = request.POST.get("estado_compra", "recibida")
    observaciones = request.POST.get("observaciones", "")
    numero_factura = request.POST.get("numero_factura", "")

    try:
     
        subtotal_compra = sum(Decimal(str(item['subtotal'])) for item in carrito)
        iva_compra = subtotal_compra * Decimal("0.19") 
        total_compra = subtotal_compra + iva_compra
        
        
        compra = Compra_proveedor.objects.create(
            numero_factura_compra=numero_factura if numero_factura else f"CMP{Compra_proveedor.objects.count()+1}",
            subtotal_compra=subtotal_compra,
            iva_compra=iva_compra,
            total_compra=total_compra,
            estado_compra=estado,
            observaciones_compra=observaciones,
            imagen_factura_compra="",
            usuarios_id_usuario=request.user,
            proveedor_idproveedor=proveedor
        )
        
     
        for item in carrito:
            producto = None
            precio_unitario = Decimal(str(item['precio']))
            cantidad = int(item['cantidad'])
            fecha_vencimiento_detalle = item['fecha_vencimiento']
            lote_detalle = item['lote']
            subtotal_linea = Decimal(str(item['subtotal']))
            
         
            if item['tipo'] == "nuevo":
                
                fecha_vencimiento_final = fecha_vencimiento_detalle
                if not fecha_vencimiento_final:
                    fecha_vencimiento_final = date.today() 
                
                categoria_obj = get_object_or_404(Categoria, id=item['categoria'])
                precio_venta = Decimal(item['precio_venta'])

                margen_calculado = Decimal('0.00')
                if precio_unitario > Decimal('0'):
                    nuevo_margen_calc = ((precio_venta - precio_unitario) / precio_unitario) * Decimal('100.00')
                    margen_calculado = nuevo_margen_calc.quantize(Decimal('0.01'))

                try:
                    producto = Producto.objects.create(
                        nombre_producto=item['nombre'],
                        descripcion_producto=item['descripcion_producto'],
                        codigo_barras=item.get('codigo_barras_nuevo', ''), 
                        registrosaniario=item['registrosaniario'],
                        precio_compra=precio_unitario,
                        precio_venta=precio_venta,
                        margen_ganancia=margen_calculado,
                        stock_actual=0, 
                        stock_minimo=int(item['stock_minimo']),
                        stock_maximo=int(item['stock_maximo']),
                        fecha_vencimiento=fecha_vencimiento_final,
                        categoria_idcategoria=categoria_obj, 
                        proveedor_idproveedor=proveedor,
                        activo=1
                    )
                except Exception as product_e:
                    
                    raise Exception(f"Fallo al crear el producto nuevo '{item['nombre']}'. Causa: {str(product_e)}")
            
            else:
             
                producto_id_from_carrito = item.get('producto_id')
                
                if not producto_id_from_carrito:
                   
                    raise Exception(f"El ID del producto existente '{item['nombre']}' es nulo o inválido en el carrito.")

                try:
                    
                    producto = Producto.objects.get(id=producto_id_from_carrito)
                except Producto.DoesNotExist:
                    raise Exception(f"El producto existente '{item['nombre']}' (ID: {producto_id_from_carrito}) ya no existe en la base de datos.")
            
            
            
            if not producto or not producto.pk:
                 raise Exception(f"El objeto Producto es nulo para '{item['nombre']}' antes de crear el detalle.")
                 
            Compra_detalle.objects.create(
                compra_idcompra=compra,
                producto_idproducto=producto, 
                cantidad=cantidad,
                precio_compra_unitario=precio_unitario,
                subtotal_linea_compra=subtotal_linea,
                lote=lote_detalle,
                fecha_vencimiento=fecha_vencimiento_detalle
            )
           
            if estado == 'recibida':
                producto.stock_actual += cantidad
                producto.precio_compra = precio_unitario 
                

                if producto.precio_venta > Decimal('0') and producto.precio_compra > Decimal('0'):
                    nuevo_margen = ((producto.precio_venta - producto.precio_compra) / producto.precio_compra) * Decimal('100.00')
                    producto.margen_ganancia = nuevo_margen.quantize(Decimal('0.01'))
                
                try:
                    producto.save() 
                except Exception as e:
                    raise Exception(f"Fallo al guardar el producto {producto.nombre_producto}: {str(e)}")


    
        request.session[carrito_key] = []
        request.session.modified = True
        
        messages.success(request, f"Compra {compra.numero_factura_compra} registrada exitosamente. Total: ${total_compra:,.2f}")
        return redirect("listar_proveedores")

    except Exception as e:
     
        messages.error(request, f'Error CRÍTICO al procesar la compra. Detalles: {str(e)}')
        return redirect('crear_compra_proveedor', idproveedor=proveedor.id)



@login_required
def crear_compra_proveedor(request, idproveedor):
    """
    Maneja la lógica para agregar productos a un carrito temporal 
    y para finalizar la compra con el proveedor.
    """
    proveedor = get_object_or_404(Proveedor, id=idproveedor)
    productos = Producto.objects.filter(activo=1)
    categorias = Categoria.objects.filter(activo=1)

    if request.method == "POST":
        if 'finalizar_compra' in request.POST:
            
            return procesar_compra_final(request, proveedor)
        elif 'agregar_producto' in request.POST:
           
            return agregar_al_carrito_compra(request, proveedor, productos, categorias)
        
        
        messages.error(request, "Acción de formulario no reconocida.")
        return redirect('crear_compra_proveedor', idproveedor=proveedor.id)


  
    carrito_key = f'carrito_compra_{proveedor.id}'
    carrito = request.session.get(carrito_key, [])
    

    subtotal_compra = sum(Decimal(str(item['subtotal'])) for item in carrito)
    iva_compra = subtotal_compra * Decimal('0.19')
    total_compra = subtotal_compra + iva_compra

    return render(request, "proveedor/crear_compra.html", {
        "proveedor": proveedor,
        "productos": productos,
        "categorias": categorias,
        "carrito": carrito,
        "subtotal_compra": subtotal_compra,
        "iva_compra": iva_compra,
        "total_compra": total_compra,
    }) 


@login_required
def compra_quitar_producto(request, idproveedor, temp_id):
    """
    Quita un producto del carrito de compra usando su ID temporal.
    """
    proveedor = get_object_or_404(Proveedor, id=idproveedor)
    carrito_key = f'carrito_compra_{proveedor.id}'
    carrito = request.session.get(carrito_key, [])
    
    
    carrito_nuevo = [item for item in carrito if item['temp_id'] != temp_id]
    
    if len(carrito_nuevo) < len(carrito):
        request.session[carrito_key] = carrito_nuevo
        request.session.modified = True
        messages.success(request, 'Producto eliminado de la lista de compra.')
    else:
        messages.error(request, 'No se encontró el producto a eliminar.')
        
    return redirect('crear_compra_proveedor', idproveedor=proveedor.id)


@login_required
def compra_limpiar_carrito(request, idproveedor):
    """
    Limpia todo el carrito de compra.
    """
    proveedor = get_object_or_404(Proveedor, id=idproveedor)
    carrito_key = f'carrito_compra_{proveedor.id}'
    
    
    request.session[carrito_key] = []
    request.session.modified = True
    
    messages.success(request, 'Lista de compra vaciada.')
    return redirect('crear_compra_proveedor', idproveedor=proveedor.id)

#DETALLE COMPRA PROVEEDOR
def detalle_compra_proveedor(request, compra_id):

    compra = get_object_or_404(Compra_proveedor, pk=compra_id)  

    detalles = Compra_detalle.objects.filter(compra_idcompra=compra)
    
    proveedor = detalles.first().producto_idproducto.proveedor_idproveedor if detalles.exists() else None
    
    context = {
        'compra': compra,
        'detalles': detalles,
        'proveedor': proveedor,
    }
    
    return render(request, 'proveedor/detalle_compra_proveedor.html', context)

def listar_compras_proveedor(request):
    compras = Compra_proveedor.objects.all().order_by('-fecha_compra')
    
    context = {
        'compras': compras,
    }
    
    return render(request, 'proveedor/listar_compras.html', context)

#PROVEEDOR

def listar_proveedores(request):
    usuario = request.user  # este es el usuario logueado

    proveedores = Proveedor.objects.all()
    return render(request, 'proveedor/listar_prov.html', {
        'proveedores': proveedores,
        'usuario': usuario
    })


@admin_required
def registrar_proveedor(request):
    # ⬇️ Se obtiene el usuario logueado para poder mostrar su info en la plantilla
    usuario = request.user  
    
    # ⬇️ Se calcula si el usuario es administrador (para mostrar el badge)
    es_admin = request.user.tipo_usu == 'administrador'

    if request.method == 'POST':
        nombre = request.POST.get('nombre_proveedor')
        correo = request.POST.get('correo_proveedor')
        telefono = request.POST.get('telefono')
        direccion = request.POST.get('direccion')
        nit = request.POST.get('nit')
        contacto_nombre = request.POST.get('contacto_nombre')
        contacto_telefono = request.POST.get('contacto_telefono')
        activo = request.POST.get('activo')
        
        if not nombre or not correo or not telefono or not direccion or not nit:
            messages.error(request, "Todos los campos obligatorios deben llenarse.")
            return redirect('registrar_proveedor')
        
        proveedor = Proveedor(
            nombre_proveedor=nombre,
            correo_proveedor=correo,
            telefono=telefono,
            direccion=direccion,
            nit=nit,
            contacto_nombre=contacto_nombre,
            contacto_telefono=contacto_telefono,
            activo=1 if activo == 'True' else 0
        )
        proveedor.save()

        return redirect('listar_proveedores')

    # ⬇️ Se envían las variables al contexto para que la plantilla pueda mostrar el badge
    return render(request, 'proveedor/registrarprov.html', {
        'usuario': usuario,
        'es_admin': es_admin
    })


@admin_required
def editar_proveedor(request, id):
    proveedor = get_object_or_404(Proveedor, id=id)

    if request.method == 'POST':
        proveedor.nombre_proveedor = request.POST['nombre_proveedor']
        proveedor.correo_proveedor = request.POST['correo_proveedor']
        proveedor.telefono = request.POST['telefono']
        proveedor.direccion = request.POST['direccion']
        proveedor.nit = request.POST['nit']
        proveedor.contacto_nombre = request.POST['contacto_nombre']
        proveedor.contacto_telefono = request.POST['contacto_telefono']
        proveedor.activo = 1 if request.POST.get('activo') == 'True' else 0
        proveedor.save()
        return redirect('listar_proveedores')

    usuario = request.user                     # <-- usuario logueado
    es_admin = request.user.tipo_usu == 'administrador'  # <-- bandera del rol

    return render(request, 'proveedor/editar_proveedor.html', {
        'proveedor': proveedor,
        'usuario': usuario,      
        'es_admin': es_admin     
    })


@admin_required
def eliminar_proveedor(request, id):
    proveedor = get_object_or_404(Proveedor, id=id)
    proveedor.delete()
    return redirect('listar_proveedores')


#LOGIN - AUTENTICACIÓN 

def login_view(request):
    """
    Vista para login de usuarios - SIN usar django.contrib.auth.login()
    """
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            correo = form.cleaned_data['correo']
            password = form.cleaned_data['password']
            
            # Autenticar usando el backend personalizado
            user = authenticate(request, correo=correo, password=password)
            
            if user is not None:
                # Crear sesión manualmente sin disparar señales
                request.session['_auth_user_id'] = user.pk
                request.session['_auth_user_backend'] = 'Sgiev.backends.UsuariosBackend'
                request.session.save()
                
                # Mensaje de bienvenida
                messages.success(request, f'Bienvenido {user.p_nombre} {user.p_apellido}')
                
                # Redirigir al dashboard
                return redirect('dashboard')
            else:
                messages.error(request, 'Credenciales inválidas')
    else:
        form = LoginForm()
    
    return render(request, 'login.html', {'form': form})


def logout_view(request):
    """
    Vista para cerrar sesión
    """
    logout(request)
    messages.success(request, 'Sesión cerrada exitosamente')
    return redirect('index')


# ===== VISTAS DE USUARIOS (CRUD) =====

@admin_required
def usuarios_listar(request):
    """
    Lista todos los usuarios con paginación y búsqueda
    """
    # Obtener parámetro de búsqueda
    search = request.GET.get('search', '')
    
    # Filtrar usuarios
    if search:
        usuarios = Usuarios.objects.filter(
            models.Q(p_nombre__icontains=search) |
            models.Q(p_apellido__icontains=search) |
            models.Q(correo__icontains=search) |
            models.Q(num_identificacion__icontains=search)
        ).order_by('-fecha_registro')
    else:
        usuarios = Usuarios.objects.all().order_by('-fecha_registro')
    
    # Paginación
    paginator = Paginator(usuarios, 10)  # 10 usuarios por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'usuario': request.user
    }
    
    return render(request, 'usuarios/listar.html', context)


@admin_required
def usuarios_crear(request):
    """
    Crear un nuevo usuario
    """
    if request.method == 'POST':
        form = UsuarioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Usuario creado exitosamente')
            return redirect('usuarios_listar')
    else:
        form = UsuarioForm()
    
    context = {
        'form': form,
        'titulo': 'Crear Usuario',
        'usuario': request.user
    }
    
    return render(request, 'usuarios/crear.html', context)


@admin_required
def usuarios_editar(request, id):
    """
    Editar un usuario existente
    """
    usuario = get_object_or_404(Usuarios, pk=id)
    
    if request.method == 'POST':
        form = UsuarioForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Usuario actualizado exitosamente')
            return redirect('usuarios_listar')
    else:
        form = UsuarioForm(instance=usuario)
    
    context = {
        'form': form,
        'titulo': 'Editar Usuario',
        'usuario': request.user,
        'usuario_editando': usuario
    }
    
    return render(request, 'usuarios/editar.html', context)


@admin_required
def usuarios_eliminar(request, id):
    """
    Eliminar (desactivar) un usuario
    """
    usuario = get_object_or_404(Usuarios, pk=id)
    
    # No permitir eliminar al usuario actual
    if usuario.id == request.user.id:
        messages.error(request, 'No puedes eliminar tu propio usuario')
        return redirect('usuarios_listar')
    
    # Desactivar en lugar de eliminar
    usuario.activo = 0
    usuario.save()
    
    messages.success(request, f'Usuario {usuario.nombre_completo} desactivado exitosamente')
    return redirect('usuarios_listar')


@admin_required
def usuarios_detalle(request, id):
    """
    Ver detalles de un usuario
    """
    usuario_detalle = get_object_or_404(Usuarios, pk=id)
    
    context = {
        'usuario_detalle': usuario_detalle,
        'usuario': request.user
    }
    
    return render(request, 'usuarios/detalle.html', context)


# ===== VISTAS DE VENTAS =====
@login_required(login_url='login')
def ventas_listar(request):
    """
    Lista todas las ventas con filtros
    - Administrador: ve todas las ventas
    - Operario: solo ve sus propias ventas
    """
    # Filtros
    search = request.GET.get('search', '')
    estado = request.GET.get('estado', '')
    metodo = request.GET.get('metodo', '')
    
    # Query base según el rol
    if request.user.tipo_usu == 'administrador':
        # Admin ve todas las ventas
        ventas = Venta.objects.all().select_related('usuarios_id_usuario').order_by('-fecha_factura')
    else:
        # Operario solo ve sus ventas
        ventas = Venta.objects.filter(usuarios_id_usuario=request.user).select_related('usuarios_id_usuario').order_by('-fecha_factura')
    
    # Aplicar filtros
    if search:
        ventas = ventas.filter(
            models.Q(numero_factura__icontains=search) |
            models.Q(usuarios_id_usuario__p_nombre__icontains=search) |
            models.Q(usuarios_id_usuario__p_apellido__icontains=search)
        )
    
    if estado:
        ventas = ventas.filter(estado_pago=estado)
    
    if metodo:
        ventas = ventas.filter(metodo_pago=metodo)
    
    # Paginación
    paginator = Paginator(ventas, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # ===== AGREGAR INFORMACIÓN DE ENVÍOS PARA CADA VENTA =====
    # Crear un diccionario con los envíos asociados a cada venta
    ventas_ids = [venta.id for venta in page_obj]
    envios_dict = {}
    
    if ventas_ids:
        envios = Envio.objects.filter(venta_idfactura__in=ventas_ids).select_related('venta_idfactura')
        for envio in envios:
            envios_dict[envio.venta_idfactura.id] = envio
    
    # Agregar el envío a cada venta en el page_obj
    for venta in page_obj:
        venta.envio_asociado = envios_dict.get(venta.id, None)
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'estado': estado,
        'metodo': metodo,
        'usuario': request.user,
        'es_admin': request.user.tipo_usu == 'administrador'
    }
    
    return render(request, 'ventas/listar.html', context)

@login_required(login_url='login')
def ventas_crear(request):
    """
    Crear nueva venta con sistema de carrito
    """
    if request.method == 'POST':
        if 'finalizar_venta' in request.POST:
            # Finalizar venta
            return procesar_venta(request)
        else:
            # Agregar producto al carrito
            return agregar_al_carrito(request)
    
    # GET - Mostrar formulario
    venta_form = VentaForm()
    producto_form = AgregarProductoForm()
    
    # Obtener carrito de la sesión
    carrito = request.session.get('carrito_venta', [])
    
    # Calcular totales
    subtotal = sum(Decimal(str(item['subtotal'])) for item in carrito)
    
    context = {
        'venta_form': venta_form,
        'producto_form': producto_form,
        'carrito': carrito,
        'subtotal': subtotal,
        'usuario': request.user,
        'es_admin': request.user.tipo_usu == 'administrador'
    }
    
    return render(request, 'ventas/crear.html', context)


def agregar_al_carrito(request):
    """
    Agrega un producto al carrito de la sesión
    """
    producto_form = AgregarProductoForm(request.POST)
    
    if producto_form.is_valid():
        producto = producto_form.cleaned_data['producto']
        cantidad = producto_form.cleaned_data['cantidad']
        
        # Obtener o crear carrito
        carrito = request.session.get('carrito_venta', [])
        
        # Verificar si el producto ya está en el carrito
        producto_existente = False
        for item in carrito:
            if item['producto_id'] == producto.id:
                # Actualizar cantidad
                nueva_cantidad = item['cantidad'] + cantidad
                if nueva_cantidad <= producto.stock_actual:
                    item['cantidad'] = nueva_cantidad
                    item['subtotal'] = float(producto.precio_venta * nueva_cantidad)
                    item['stock_disponible'] = producto.stock_actual  # Actualizar stock
                    producto_existente = True
                    messages.success(request, f'Cantidad actualizada: {producto.nombre_producto}')
                else:
                    messages.error(request, f'Stock insuficiente para {producto.nombre_producto}')
                    return redirect('ventas_crear')
                break
        
        if not producto_existente:
            # Agregar nuevo producto
            carrito.append({
                'producto_id': producto.id,
                'nombre': producto.nombre_producto,
                'precio': float(producto.precio_venta),
                'cantidad': cantidad,
                'subtotal': float(producto.precio_venta * cantidad),
                'stock_disponible': producto.stock_actual,
                'stock_minimo': producto.stock_minimo  # AGREGAR ESTO
            })
            messages.success(request, f'Producto agregado: {producto.nombre_producto}')
        
        # Guardar carrito en sesión
        request.session['carrito_venta'] = carrito
        request.session.modified = True
    else:
        for error in producto_form.errors.values():
            messages.error(request, error)
    
    return redirect('ventas_crear')


@login_required(login_url='login')
def ventas_quitar_producto(request, producto_id):
    """
    Quita un producto del carrito
    """
    carrito = request.session.get('carrito_venta', [])
    carrito = [item for item in carrito if item['producto_id'] != producto_id]
    request.session['carrito_venta'] = carrito
    request.session.modified = True
    
    messages.success(request, 'Producto eliminado del carrito')
    return redirect('ventas_crear')


@login_required(login_url='login')
def ventas_limpiar_carrito(request):
    """
    Limpia todo el carrito
    """
    request.session['carrito_venta'] = []
    request.session.modified = True
    
    messages.success(request, 'Carrito vaciado')
    return redirect('ventas_crear')

@transaction.atomic
def procesar_venta(request):
    """
    Procesa la venta final con validación de abono mínimo y estado automático
    """
    venta_form = VentaForm(request.POST)
    carrito = request.session.get('carrito_venta', [])
    
    if not carrito:
        messages.error(request, 'El carrito está vacío')
        return redirect('ventas_crear')
    
    if venta_form.is_valid():
        try:
            # Calcular totales
            subtotal = sum(Decimal(str(item['subtotal'])) for item in carrito)
            descuento = venta_form.cleaned_data['descuento'] or Decimal('0')
            iva = (subtotal - descuento) * Decimal('0.19')
            valor_total = subtotal - descuento + iva
            abono = venta_form.cleaned_data['abono'] or Decimal('0')
            
            # ===== VALIDACIÓN DE ABONO MÍNIMO =====
            abono_minimo = valor_total * Decimal('0.10')  # 10% del total
            
            if abono > 0 and abono < abono_minimo:
                messages.error(
                    request, 
                    f'El abono mínimo debe ser el 10% del total (${abono_minimo:,.0f}). '
                    f'Si no desea abonar, deje el campo en 0.'
                )
                return redirect('ventas_crear')
            
            # ===== CALCULAR ESTADO DE PAGO AUTOMÁTICAMENTE =====
            if abono == 0:
                estado_pago = 'pendiente'
            elif abono >= valor_total:
                estado_pago = 'pagado'
                abono = valor_total  # Ajustar si pagó de más
            else:
                estado_pago = 'parcial'
            
            saldo_pendiente = valor_total - abono
            
            # Crear venta
            venta = venta_form.save(commit=False)
            venta.subtotal = subtotal
            venta.iva = iva
            venta.valor_total = valor_total
            venta.abono = abono
            venta.saldo_pendiente = saldo_pendiente
            venta.estado_pago = estado_pago  # Estado automático
            venta.usuarios_id_usuario = request.user
            venta.imagen_recibo = ''
            venta.save()
            
            # Lista para productos con stock bajo
            productos_stock_bajo = []
            
            # Procesar cada producto del carrito
            for item in carrito:
                producto = Producto.objects.get(id=item['producto_id'])
                cantidad = item['cantidad']
                
                if producto.stock_actual < cantidad:
                    raise Exception(f'Stock insuficiente para {producto.nombre_producto}')
                
                # Crear detalle de venta
                Venta_has_producto.objects.create(
                    venta_idfactura=venta,
                    producto_idproducto=producto,
                    cantidad=cantidad,
                    valor_unitario=producto.precio_venta,
                    subtotal_linea=Decimal(str(item['subtotal']))
                )
                
                # Actualizar stock
                stock_anterior = producto.stock_actual
                producto.stock_actual -= cantidad
                producto.save()
                
                # Verificar stock bajo
                if producto.stock_actual <= producto.stock_minimo:
                    productos_stock_bajo.append({
                        'nombre': producto.nombre_producto,
                        'stock_actual': producto.stock_actual,
                        'stock_minimo': producto.stock_minimo
                    })
                
                # Registrar movimiento de inventario
                Movimiento_inventario.objects.create(
                    tipo_movimiento='venta',
                    cantidad=cantidad,
                    stock_anterior=stock_anterior,
                    stock_nuevo=producto.stock_actual,
                    precio_unitario=producto.precio_venta,
                    valor_total=Decimal(str(item['subtotal'])),
                    referencia_id=venta.id,
                    tipo_referencia='venta',
                    observaciones=f'Venta #{venta.numero_factura}',
                    imagen_comprobante='',
                    usuarios_id_usuario=request.user,
                    producto_idproducto=producto
                )
            
            # Limpiar carrito
            request.session['carrito_venta'] = []
            request.session.modified = True
            
            # Mensajes de éxito
            estado_texto = {
                'pendiente': 'PENDIENTE',
                'parcial': 'PARCIAL',
                'pagado': 'PAGADO'
            }
            messages.success(
                request, 
                f'Venta {venta.numero_factura} registrada exitosamente. '
                f'Estado: {estado_texto[estado_pago]}'
            )
            
            # Alertar sobre productos con stock bajo
            if productos_stock_bajo:
                for prod in productos_stock_bajo:
                    messages.warning(
                        request, 
                        f'⚠️ ALERTA: {prod["nombre"]} tiene stock bajo. '
                        f'Actual: {prod["stock_actual"]} | Mínimo: {prod["stock_minimo"]}'
                    )
            
            return redirect('ventas_detalle', id=venta.id)
            
        except Exception as e:
            messages.error(request, f'Error al procesar la venta: {str(e)}')
            return redirect('ventas_crear')
    else:
        for error in venta_form.errors.values():
            messages.error(request, error)
        return redirect('ventas_crear')


@login_required(login_url='login')
def ventas_detalle(request, id):
    """
    Ver detalles completos de una venta (factura)
    """
    venta = get_object_or_404(Venta, pk=id)
    productos = Venta_has_producto.objects.filter(venta_idfactura=venta)
    
    context = {
        'venta': venta,
        'productos': productos,
        'usuario': request.user,
        'es_admin': request.user.tipo_usu == 'administrador'
    }
    
    return render(request, 'ventas/detalle.html', context)


@login_required(login_url='login')
def obtener_precio_producto(request, producto_id):
    """
    API para obtener información de un producto (AJAX)
    """
    try:
        producto = Producto.objects.get(id=producto_id, activo=1)
        data = {
            'precio': float(producto.precio_venta),
            'stock': producto.stock_actual,
            'nombre': producto.nombre_producto
        }
        return JsonResponse(data)
    except Producto.DoesNotExist:
        return JsonResponse({'error': 'Producto no encontrado'}, status=404)

@admin_required
def ventas_editar_estado(request, id):
    """
    Permite al administrador editar solo el estado de pago de una venta
    """
    venta = get_object_or_404(Venta, pk=id)
    
    if request.method == 'POST':
        form = EditarEstadoVentaForm(request.POST, instance=venta)
        if form.is_valid():
            form.save()
            messages.success(request, f'Estado de la venta {venta.numero_factura} actualizado')
            return redirect('ventas_detalle', id=venta.id)
    else:
        form = EditarEstadoVentaForm(instance=venta)
    
    context = {
        'form': form,
        'venta': venta,
        'usuario': request.user,
        'es_admin': True
    }
    
    return render(request, 'ventas/editar_estado.html', context)


@login_required(login_url='login')
def ventas_generar_pdf(request, id):
    """
    Genera un PDF de la factura de venta
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from django.http import HttpResponse
    from io import BytesIO
    
    # Obtener la venta
    venta = get_object_or_404(Venta, pk=id)
    productos = Venta_has_producto.objects.filter(venta_idfactura=venta)
    
    # Crear el objeto HttpResponse con el tipo de contenido PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Factura_{venta.numero_factura}.pdf"'
    
    # Crear el PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#3d862e'),
        alignment=TA_CENTER
    )
    
    # Título
    elements.append(Paragraph("ROMAR NATURAL", title_style))
    elements.append(Paragraph("NIT: 52101085", styles['Normal']))
    elements.append(Paragraph("Teléfono: 3053615676", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Información de la factura
    elements.append(Paragraph(f"<b>FACTURA: {venta.numero_factura}</b>", styles['Heading2']))
    elements.append(Paragraph(f"Fecha: {venta.fecha_factura.strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Paragraph(f"Vendedor: {venta.usuarios_id_usuario.nombre_completo}", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Tabla de productos
    data = [['#', 'Producto', 'Cant.', 'Precio Unit.', 'Subtotal']]
    
    for idx, item in enumerate(productos, 1):
        data.append([
            str(idx),
            item.producto_idproducto.nombre_producto[:30],
            str(item.cantidad),
            f"${item.valor_unitario:,.0f}",
            f"${item.subtotal_linea:,.0f}"
        ])
    
    # Crear tabla
    table = Table(data, colWidths=[0.5*inch, 3*inch, 0.8*inch, 1.2*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3d862e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Totales
    totales_data = [
        ['Subtotal:', f"${venta.subtotal:,.0f}"],
        ['Descuento:', f"-${venta.descuento:,.0f}"],
        ['IVA (19%):', f"${venta.iva:,.0f}"],
        ['<b>TOTAL:</b>', f"<b>${venta.valor_total:,.0f}</b>"]
    ]
    
    if venta.abono > 0:
        totales_data.append(['Abono:', f"${venta.abono:,.0f}"])
        totales_data.append(['<b>Saldo Pendiente:</b>', f"<b>${venta.saldo_pendiente:,.0f}</b>"])
    
    totales_table = Table(totales_data, colWidths=[3*inch, 2*inch])
    totales_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 14),
        ('LINEABOVE', (0, -2), (-1, -2), 2, colors.black),
    ]))
    
    elements.append(totales_table)
    elements.append(Spacer(1, 0.5*inch))
    
    # Información adicional
    elements.append(Paragraph(f"<b>Método de Pago:</b> {venta.get_metodo_pago_display()}", styles['Normal']))
    elements.append(Paragraph(f"<b>Estado:</b> {venta.get_estado_pago_display()}", styles['Normal']))
    
    if venta.observaciones:
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph(f"<b>Observaciones:</b> {venta.observaciones}", styles['Normal']))
    
    elements.append(Spacer(1, 0.5*inch))
    elements.append(Paragraph("Gracias por su compra", styles['Normal']))
    
    # Construir PDF
    doc.build(elements)
    
    # Obtener el valor del buffer y escribirlo en la respuesta
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response
@admin_required
@transaction.atomic
def ventas_eliminar(request, id):
    """
    Elimina una venta y devuelve el stock al inventario
    Solo disponible para administradores
    """
    venta = get_object_or_404(Venta, pk=id)
    
    try:
        # Obtener todos los productos de la venta
        productos_venta = Venta_has_producto.objects.filter(venta_idfactura=venta)
        
        # Revertir el stock de cada producto
        for item in productos_venta:
            producto = item.producto_idproducto
            cantidad = item.cantidad
            
            # Devolver el stock
            stock_anterior = producto.stock_actual
            producto.stock_actual += cantidad
            producto.save()
            
            # Registrar movimiento de reversión
            Movimiento_inventario.objects.create(
                tipo_movimiento='ajuste',
                cantidad=cantidad,
                stock_anterior=stock_anterior,
                stock_nuevo=producto.stock_actual,
                precio_unitario=producto.precio_venta,
                valor_total=item.subtotal_linea,
                referencia_id=venta.id,
                tipo_referencia='ajuste',
                observaciones=f'Reversión por eliminación de venta {venta.numero_factura}',
                imagen_comprobante='',
                usuarios_id_usuario=request.user,
                producto_idproducto=producto
            )
        
        # Guardar información antes de eliminar
        numero_factura = venta.numero_factura
        
        # Eliminar la venta (esto eliminará también los detalles por CASCADE)
        venta.delete()
        
        messages.success(
            request, 
            f'Venta {numero_factura} eliminada exitosamente. El stock ha sido devuelto al inventario.'
        )
        
    except Exception as e:
        messages.error(request, f'Error al eliminar la venta: {str(e)}')
    
    return redirect('ventas_listar')

# ===== VISTAS DE MENSAJERÍA (SOLO ADMIN) =====

@admin_required
def mensajeria_listar(request):
    """
    Lista todas las empresas de mensajería (solo admin)
    """
    search = request.GET.get('search', '')
    
    if search:
        mensajerias = Mensajeria.objects.filter(
            models.Q(nombre_mensajeria__icontains=search) |
            models.Q(cobertura__icontains=search)
        ).order_by('nombre_mensajeria')
    else:
        mensajerias = Mensajeria.objects.all().order_by('nombre_mensajeria')
    
    paginator = Paginator(mensajerias, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'usuario': request.user,
        'es_admin': True
    }
    
    return render(request, 'mensajeria/listar.html', context)


@admin_required
def mensajeria_crear(request):
    """
    Crear nueva empresa de mensajería (solo admin)
    """
    if request.method == 'POST':
        form = MensajeriaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Empresa de mensajería creada exitosamente')
            return redirect('mensajeria_listar')
    else:
        form = MensajeriaForm()
    
    context = {
        'form': form,
        'titulo': 'Nueva Empresa de Mensajería',
        'usuario': request.user,
        'es_admin': True
    }
    
    return render(request, 'mensajeria/crear.html', context)


@admin_required
def mensajeria_editar(request, id):
    """
    Editar empresa de mensajería (solo admin)
    """
    mensajeria = get_object_or_404(Mensajeria, pk=id)
    
    if request.method == 'POST':
        form = MensajeriaForm(request.POST, instance=mensajeria)
        if form.is_valid():
            form.save()
            messages.success(request, 'Empresa de mensajería actualizada exitosamente')
            return redirect('mensajeria_listar')
    else:
        form = MensajeriaForm(instance=mensajeria)
    
    context = {
        'form': form,
        'titulo': 'Editar Empresa de Mensajería',
        'mensajeria': mensajeria,
        'usuario': request.user,
        'es_admin': True
    }
    
    return render(request, 'mensajeria/editar.html', context)


@admin_required
def mensajeria_eliminar(request, id):
    """
    Eliminar empresa de mensajería (solo admin)
    """
    mensajeria = get_object_or_404(Mensajeria, pk=id)
    nombre = mensajeria.nombre_mensajeria
    mensajeria.delete()
    
    messages.success(request, f'Empresa {nombre} eliminada exitosamente')
    return redirect('mensajeria_listar')


# ===== VISTAS DE ENVÍOS =====

@login_required(login_url='login')
def envios_listar(request):
    """
    Lista todos los envíos
    Admin: ve todos | Operario: ve los suyos
    + Muestra ventas sin envío asignado
    """
    search = request.GET.get('search', '')
    estado = request.GET.get('estado', '')
    search_pendientes = request.GET.get('search_pendientes', '')
    
    # Filtrar según rol
    if request.user.tipo_usu == 'administrador':
        envios = Envio.objects.all().order_by('-fecha_envio')
        ventas_query = Venta.objects.all()
    else:
        envios = Envio.objects.filter(usuarios_id_usuario=request.user).order_by('-fecha_envio')
        ventas_query = Venta.objects.filter(usuarios_id_usuario=request.user)
    
    # Aplicar filtros a envíos
    if search:
        envios = envios.filter(
            models.Q(venta_idfactura__numero_factura__icontains=search) |
            models.Q(direccion_envio__icontains=search) |
            models.Q(fk_mensajeria__nombre_mensajeria__icontains=search)
        )
    
    if estado:
        envios = envios.filter(estado_envio=estado)
    
    # Paginación de envíos
    paginator = Paginator(envios, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # ===== OBTENER VENTAS SIN ENVÍO ASIGNADO =====
    ventas_con_envio = Envio.objects.values_list('venta_idfactura', flat=True)
    ventas_sin_envio = ventas_query.exclude(id__in=ventas_con_envio).select_related('usuarios_id_usuario').order_by('-fecha_factura')
    
    # Filtrar ventas pendientes si hay búsqueda
    if search_pendientes:
        ventas_sin_envio = ventas_sin_envio.filter(numero_factura__icontains=search_pendientes)
    
    # Paginación de ventas sin envío
    paginator_pendientes = Paginator(ventas_sin_envio, 5)
    page_pendientes = request.GET.get('page_pendientes')
    ventas_pendientes_obj = paginator_pendientes.get_page(page_pendientes)
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'estado': estado,
        'ventas_pendientes_obj': ventas_pendientes_obj,
        'search_pendientes': search_pendientes,
        'total_pendientes': ventas_sin_envio.count(),
        'usuario': request.user,
        'es_admin': request.user.tipo_usu == 'administrador'
    }
    
    return render(request, 'envios/listar.html', context)



@login_required(login_url='login')
def envios_crear(request):
    """
    Crear nuevo envío
    """
    # Obtener venta_id si viene por parámetro GET
    venta_id = request.GET.get('venta_id', None)
    
    if request.method == 'POST':
        form = EnvioForm(request.POST)
        if form.is_valid():
            envio = form.save(commit=False)
            envio.usuarios_id_usuario = request.user
            envio.save()
            messages.success(request, f'Envío registrado exitosamente para la venta {envio.venta_idfactura.numero_factura}')
            return redirect('envios_detalle', id=envio.id)
    else:
        # Si viene venta_id, preseleccionar esa venta en el formulario
        if venta_id:
            try:
                venta = Venta.objects.get(id=venta_id)
                # Verificar que no tenga envío ya asociado
                if Envio.objects.filter(venta_idfactura=venta).exists():
                    messages.warning(request, f'La venta {venta.numero_factura} ya tiene un envío asociado.')
                    return redirect('ventas_listar')
                
                form = EnvioForm(initial={'venta_idfactura': venta})
            except Venta.DoesNotExist:
                messages.error(request, 'Venta no encontrada')
                form = EnvioForm()
        else:
            form = EnvioForm()
    
    context = {
        'form': form,
        'titulo': 'Registrar Nuevo Envío',
        'usuario': request.user,
        'es_admin': request.user.tipo_usu == 'administrador'
    }
    
    return render(request, 'envios/crear.html', context)


@login_required(login_url='login')
def envios_editar(request, id):
    """
    Editar envío existente
    Admin: puede editar todo
    Operario: solo puede editar estado y novedades
    """
    envio = get_object_or_404(Envio, pk=id)
    es_admin = request.user.tipo_usu == 'administrador'
    
    if request.method == 'POST':
        # Usar formulario según el rol
        if es_admin:
            form = EnvioForm(request.POST, instance=envio)
        else:
            form = EnvioEditarOperarioForm(request.POST, instance=envio)
        
        if form.is_valid():
            form.save()
            messages.success(request, 'Envío actualizado exitosamente')
            return redirect('envios_detalle', id=envio.id)
    else:
        # Cargar formulario según el rol CON DATOS EXISTENTES
        if es_admin:
            form = EnvioForm(instance=envio)  # Formulario completo con todos los datos
        else:
            form = EnvioEditarOperarioForm(instance=envio)  # Formulario limitado
    
    context = {
        'form': form,
        'titulo': 'Editar Envío',
        'envio': envio,
        'usuario': request.user,
        'es_admin': es_admin,
        'solo_estado': not es_admin  # Variable para el template
    }
    
    return render(request, 'envios/editar.html', context)

@admin_required
def envios_eliminar(request, id):
    """
    Eliminar envío (solo admin)
    """
    envio = get_object_or_404(Envio, pk=id)
    factura = envio.venta_idfactura.numero_factura
    envio.delete()
    
    messages.success(request, f'Envío de la factura {factura} eliminado exitosamente')
    return redirect('envios_listar')


@login_required(login_url='login')
def envios_detalle(request, id):
    """
    Ver detalles de un envío
    """
    envio = get_object_or_404(Envio, pk=id)
    
    context = {
        'envio': envio,
        'usuario': request.user,
        'es_admin': request.user.tipo_usu == 'administrador'
    }
    
    return render(request, 'envios/detalle.html', context)
